from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

from .db import get_project, list_documents_with_latest
from .logo_store import list_logos
from .templates_store import load_templates, resolve_abs_path

TEMPLATE_DOC_ID = "MI-DT-PJ-007"
HEADERS = ["Rev", "Document No.", "Document Type", "File Type", "Description", "Status", "Comments"]
HEADER_ROW = 9
DATA_START_ROW = 10


def _safe_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def reports_dir_for_db(db_path: Path) -> Path:
    db_path = Path(db_path)
    base = db_path.parent
    if base.name.startswith("."):
        base = base.parent
    out = base / "Reports"
    out.mkdir(parents=True, exist_ok=True)
    return out


def _rows_for_project(db_path: Path, project_id: int) -> List[Dict[str, Any]]:
    return list_documents_with_latest(Path(db_path), int(project_id), state="active") or []


def _is_drawing_row(row: Dict[str, Any]) -> bool:
    hay = f"{_safe_str(row.get('doc_type'))} {_safe_str(row.get('file_type'))}".upper()
    return ("DRAWING" in hay) or ("DWG" in hay)


def export_drawing_index_data_link_xlsx(db_path: Path, project_id: int, out_path: Path) -> Path:
    """Export Doc ID + Description for rows where Type/File Type contains DRAWING or DWG."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo

    db_path = Path(db_path)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    rows = [r for r in _rows_for_project(db_path, int(project_id)) if _is_drawing_row(r)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Drawing Index Data Link"

    header_fill = PatternFill("solid", fgColor="007F4D")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D0D6DF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in rows:
        ws.append([_safe_str(row.get("doc_id")), _safe_str(row.get("description"))])

    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in r:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 80

    wb.save(str(out_path))
    return out_path


def _find_native_register_template() -> Path:
    items = load_templates() or []
    for item in items:
        if _safe_str(item.get("doc_id")).upper() == TEMPLATE_DOC_ID:
            path = Path(item.get("abs_path") or "")
            if not path.exists():
                try:
                    path = resolve_abs_path(item)
                except Exception:
                    pass
            if path.exists():
                return path
            raise RuntimeError(f"Template {TEMPLATE_DOC_ID} is configured but does not exist:\n{path}")
    raise RuntimeError(
        f"Template {TEMPLATE_DOC_ID} was not found in Templates.\n\n"
        "Add the Excel register template in Project Settings / Templates first."
    )


def _native_register_output_path(db_path: Path, template_path: Optional[Path] = None) -> Path:
    suffix = ".xlsx"
    if template_path and template_path.suffix.lower() == ".xlsm":
        suffix = ".xlsm"
    return reports_dir_for_db(Path(db_path)) / f"{Path(db_path).stem} (Native){suffix}"


def _register_values(rows: Sequence[Dict[str, Any]]) -> List[List[str]]:
    vals: List[List[str]] = []
    for row in rows:
        vals.append([
            _safe_str(row.get("latest_rev")),
            _safe_str(row.get("doc_id")),
            _safe_str(row.get("doc_type")),
            _safe_str(row.get("file_type")),
            _safe_str(row.get("description")),
            _safe_str(row.get("status")),
            _safe_str(row.get("comments")),
        ])
    return vals


def _header_project_values(project: Dict[str, Any]) -> Dict[str, str]:
    return {
        "doc_id": TEMPLATE_DOC_ID,
        "project_code": _safe_str(project.get("project_code")),
        "client": _safe_str(project.get("client_company")) or _safe_str(project.get("client_reference")),
        "end_user": _safe_str(project.get("end_user")),
    }


def _try_xlwings_export(dest_path: Path, db_path: Path, project: Dict[str, Any], rows: Sequence[Dict[str, Any]]) -> bool:
    try:
        import xlwings as xw  # type: ignore
    except Exception:
        return False

    app = None
    book = None
    try:
        app = xw.App(visible=False, add_book=False)
        book = xw.Book(str(dest_path))
        sheet_names = [s.name for s in book.sheets]
        header_sht = book.sheets["Cover Sheet"] if "Cover Sheet" in sheet_names else book.sheets[0]
        data_sht = book.sheets["MI Documents"] if "MI Documents" in sheet_names else header_sht

        hv = _header_project_values(project)
        header_sht.range("I4").value = hv["doc_id"]
        header_sht.range("I6").value = hv["project_code"]
        header_sht.range("I8").value = hv["client"]
        header_sht.range("I10").value = hv["end_user"]

        try:
            header_sht.range("A7").value = None
        except Exception:
            pass

        # Remove old injected client logos, then add current project logos.
        try:
            for pic in list(header_sht.pictures):
                if str(getattr(pic, "name", "")).startswith("ClientLogo_"):
                    pic.delete()
        except Exception:
            pass

        logos = list_logos(Path(db_path))
        if logos:
            try:
                zone = header_sht.range("A7").merge_area
            except Exception:
                zone = header_sht.range("A7")
            max_logos = min(len(logos), 3)
            gutter = 6.0
            slot_w = (zone.width - gutter * (max_logos - 1)) / max(1, max_logos)
            slot_h = zone.height
            for i, logo in enumerate(logos[:max_logos]):
                try:
                    left_slot = zone.left + i * (slot_w + gutter)
                    pic = header_sht.pictures.add(str(logo), name=f"ClientLogo_{i+1}", left=left_slot, top=zone.top)
                    w0, h0 = float(pic.width), float(pic.height)
                    if w0 > 0 and h0 > 0:
                        scale = min(slot_w / w0, slot_h / h0, 1.0)
                        pic.width = w0 * scale
                        pic.height = h0 * scale
                        pic.left = left_slot + (slot_w - pic.width) / 2.0
                        pic.top = zone.top + (slot_h - pic.height) / 2.0
                except Exception:
                    continue

        values = _register_values(rows)
        # Keep headers fixed at row 9; data starts at row 10.
        data_sht.range((HEADER_ROW, 1)).value = HEADERS
        clear_to = max(DATA_START_ROW + len(values) + 25, 500)
        data_sht.range(f"A{DATA_START_ROW}:G{clear_to}").clear_contents()
        if values:
            data_sht.range((DATA_START_ROW, 1)).value = values
        try:
            data_sht.range("A:G").autofit()
        except Exception:
            pass

        book.save()
        return True
    except Exception as exc:
        try:
            print(f"[excel_exports] xlwings route failed: {exc}", flush=True)
        except Exception:
            pass
        return False
    finally:
        try:
            if book is not None:
                book.close()
        except Exception:
            pass
        try:
            if app is not None:
                app.quit()
        except Exception:
            pass


def _openpyxl_export(dest_path: Path, db_path: Path, project: Dict[str, Any], rows: Sequence[Dict[str, Any]]) -> Path:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = load_workbook(str(dest_path), keep_vba=(dest_path.suffix.lower() == ".xlsm"))
    header_ws = wb["Cover Sheet"] if "Cover Sheet" in wb.sheetnames else wb.active
    data_ws = wb["MI Documents"] if "MI Documents" in wb.sheetnames else header_ws

    hv = _header_project_values(project)
    header_ws["I4"] = hv["doc_id"]
    header_ws["I6"] = hv["project_code"]
    header_ws["I8"] = hv["client"]
    header_ws["I10"] = hv["end_user"]
    try:
        header_ws["A7"].value = ""
    except Exception:
        pass

    # Add logos where openpyxl supports the image type.
    anchors = ["A7", "F7", "K7"]
    for idx, logo in enumerate(list_logos(Path(db_path))[:3]):
        if logo.suffix.lower() not in {".png", ".jpg", ".jpeg", ".bmp", ".gif"}:
            continue
        try:
            img = XLImage(str(logo))
            # keep roughly within a header band
            if img.width and img.height:
                scale = min(220 / float(img.width), 85 / float(img.height), 1.0)
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
            header_ws.add_image(img, anchors[min(idx, len(anchors) - 1)])
        except Exception:
            continue

    # Headers at row 9, register data from row 10.
    header_fill = PatternFill("solid", fgColor="007F4D")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D0D6DF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(HEADERS, start=1):
        cell = data_ws.cell(HEADER_ROW, col)
        cell.value = title
        if cell.fill.fill_type is None:
            cell.fill = header_fill
        cell.font = copy(cell.font)
        cell.font = Font(name=cell.font.name or "Arial", size=cell.font.sz or 10, bold=True,
                         italic=cell.font.italic, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Capture row-10 template style before clearing values.
    template_styles = {}
    for col in range(1, len(HEADERS) + 1):
        template_styles[col] = copy(data_ws.cell(DATA_START_ROW, col)._style)

    max_clear = max(data_ws.max_row, DATA_START_ROW + len(rows) + 25, 500)
    for r in range(DATA_START_ROW, max_clear + 1):
        for c in range(1, len(HEADERS) + 1):
            data_ws.cell(r, c).value = None

    values = _register_values(rows)
    for r_off, row_vals in enumerate(values, start=DATA_START_ROW):
        for c, value in enumerate(row_vals, start=1):
            cell = data_ws.cell(r_off, c)
            cell.value = value
            try:
                cell._style = copy(template_styles.get(c, cell._style))
            except Exception:
                pass
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    data_ws.column_dimensions["A"].width = 10
    data_ws.column_dimensions["B"].width = 36
    data_ws.column_dimensions["C"].width = 22
    data_ws.column_dimensions["D"].width = 18
    data_ws.column_dimensions["E"].width = 70
    data_ws.column_dimensions["F"].width = 24
    data_ws.column_dimensions["G"].width = 36
    try:
        data_ws.freeze_panes = f"A{DATA_START_ROW}"
        data_ws.auto_filter.ref = f"A{HEADER_ROW}:G{max(HEADER_ROW, DATA_START_ROW + len(values) - 1)}"
    except Exception:
        pass

    wb.save(str(dest_path))
    return dest_path


def export_native_register_xlsx(db_path: Path, project_id: int, out_path: Optional[Path] = None) -> Path:
    """Create the native register Excel export from template MI-DT-PJ-007."""
    db_path = Path(db_path)
    project = get_project(db_path) or {}
    if not project:
        raise RuntimeError("Project metadata not found in this database.")

    template_path = _find_native_register_template()
    dest_path = Path(out_path) if out_path else _native_register_output_path(db_path, template_path)
    dest_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy template first so both xw and openpyxl modify a generated file.
    import shutil
    shutil.copy2(str(template_path), str(dest_path))

    rows = _rows_for_project(db_path, int(project_id))

    if _try_xlwings_export(dest_path, db_path, project, rows):
        return dest_path

    return _openpyxl_export(dest_path, db_path, project, rows)
